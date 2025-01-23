from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

wb = load_workbook('data.xlsx')
ws = wb.active

firstRow = 2
lastRow = ws.max_row + 1
idx = 1

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True, color='FF0000')

for i in range(firstRow, lastRow):
  try:
    prdName  = '키즈꼬모 ' + ws.cell(i, 1).value
    driver.get('https://search.shopping.naver.com/search/all?query={}&sort=price_asc&fo=true'.format(prdName ))
    time.sleep(1)
    try:
      searchPrdInfo = driver.find_element(By.CSS_SELECTOR, '#content > div.style_content__xWg5l > div.basicList_list_basis__uNBZx > div > div:nth-child(1) > div > div.product_info_area__xxCTi > div.product_title__Mmw2K > a').text
    except:
      pass
    ws.cell(i, 2).value = driver.current_url
    if ws.cell(i, 1).value not in searchPrdInfo:
      ws.cell(i, 3).value = '확인필요'
      for row in range(1, 4):
        ws.cell(i, row).fill = fillData
      ws.cell(i, 3).alignment = fillAlignment
      ws.cell(i, 3).font = fillFont
    else:
      ws.cell(i, 3).value = ' '
    print('{} / {} / {}'.format(idx, prdName, driver.current_url))
    idx += 1
  except Exception as e:
    print(e)
  
wb.save('data_네이버 최저가 URL 추출본.xlsx')

driver.quit()